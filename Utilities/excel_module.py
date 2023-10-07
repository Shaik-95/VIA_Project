import openpyxl


class HomePageData:

    Test_data = [{"from_loc":"C","to_loc":"H","name":"Aulleem", "email":"shaik.abdulaleem.95@gmail.com", "mobile": "9030537469"}]
    

    
    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\Users\shaik\PycharmProjects\Project_VIA\Excell_files\data1.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1): 
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]
